import csv
import os
import traceback
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def save_to_csv(data, filename=None):
    """Save business data to a CSV file with specific column order."""
    if not os.path.exists("dat"):
        os.makedirs("dat")

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"dat/happy_hour_businesses_{timestamp}.csv"
    else:
        filename = os.path.join("dat", filename)

    # Determine max emails for dynamic columns
    max_emails = 1
    for row in data:
        emails = row.get("emails", [])
        max_emails = max(max_emails, len(emails))

    # Define column headers
    ordered_columns = ["name", "website"]
    for i in range(1, max_emails + 1):
        ordered_columns.append(f"email_{i}")
    ordered_columns.extend(["city", "state", "zip_code", "phone", "happy_hour"])

    # Write to CSV
    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ordered_columns)
        writer.writeheader()
        for row in data:
            # Create a new dict with desired fields
            cleaned_row = {}

            # Process fields
            for column in ordered_columns:
                if column.startswith("email_"):
                    continue  # Handle emails separately
                elif column == "happy_hour":
                    cleaned_row[column] = "Yes"
                else:
                    cleaned_row[column] = row.get(column, "")

            # Handle emails
            emails = row.get("emails", [])
            for i in range(min(max_emails, len(emails))):
                cleaned_row[f"email_{i+1}"] = emails[i]
            for i in range(len(emails), max_emails):
                cleaned_row[f"email_{i+1}"] = ""

            writer.writerow(cleaned_row)

    return filename


def create_excel_workbook(filename):
    """Create an empty Excel workbook with the given filename."""
    if not os.path.exists("dat"):
        os.makedirs("dat")
        print(f"Created directory: dat")

    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    full_path = os.path.join("dat", filename)
    print(f"Creating Excel file: {full_path}")

    wb = Workbook()
    # Keep default sheet for now
    wb.active.title = "Info"

    # Save workbook
    try:
        wb.save(full_path)
        print(f"Successfully created Excel file: {full_path}")
        return full_path
    except Exception as e:
        print(f"Error creating Excel file: {str(e)}")
        traceback.print_exc()
        return None


def save_to_excel(data, filename=None, sheet_name="Businesses"):
    """Save business data to a new Excel file with specific column order."""
    try:
        if not os.path.exists("dat"):
            os.makedirs("dat")
            print(f"Created directory: dat")

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"happy_hour_businesses_{timestamp}.xlsx"

        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        full_path = os.path.join("dat", filename)
        print(f"Saving Excel file to: {full_path}")

        # Create a new workbook
        wb = Workbook()
        ws = wb.active

        # Rename the default sheet
        if sheet_name:
            ws.title = sanitize_sheet_name(sheet_name)

        # Process data
        _write_data_to_worksheet(ws, data)

        # Save the workbook
        wb.save(full_path)
        print(f"Successfully saved Excel file: {full_path}")
        return full_path

    except Exception as e:
        print(f"Error in save_to_excel: {str(e)}")
        traceback.print_exc()
        return None


def add_sheet_to_excel(filename, data, sheet_name):
    """Add a new sheet to an existing Excel file."""
    try:
        print(f"Attempting to add sheet to Excel file: {filename}")

        # Ensure the file exists first
        if not os.path.exists(filename):
            if os.path.exists(os.path.join("dat", filename)):
                filename = os.path.join("dat", filename)
            else:
                print(f"Excel file not found: {filename}")
                print(f"Creating new file instead")
                return save_to_excel(data, filename, sheet_name)

        # Load the Excel file
        wb = load_workbook(filename)
        print(f"Successfully loaded Excel file: {filename}")

        # Create a safe sheet name
        safe_sheet_name = sanitize_sheet_name(sheet_name)

        # Handle duplicate sheet names
        count = 1
        original_name = safe_sheet_name
        while safe_sheet_name in wb.sheetnames:
            print(
                f"Sheet name '{safe_sheet_name}' already exists, creating alternate name"
            )
            safe_sheet_name = f"{original_name[:27]}_{count}"
            count += 1
            if count > 100:
                raise ValueError(f"Could not create unique sheet name for {sheet_name}")

        # Create the new sheet
        ws = wb.create_sheet(safe_sheet_name)
        print(f"Created new sheet: {safe_sheet_name}")

        # Write data to the sheet
        _write_data_to_worksheet(ws, data)

        # Save the workbook
        wb.save(filename)
        print(f"Successfully saved updated Excel file: {filename}")
        return filename

    except Exception as e:
        print(f"Error in add_sheet_to_excel: {str(e)}")
        traceback.print_exc()
        return None


def sanitize_sheet_name(sheet_name):
    """Sanitize a sheet name for Excel."""
    if not sheet_name:
        return "Sheet1"

    # Remove invalid characters
    invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
    safe_name = sheet_name
    for char in invalid_chars:
        safe_name = safe_name.replace(char, "-")

    # Limit to 31 characters (Excel limit)
    safe_name = safe_name[:31]

    return safe_name


def _write_data_to_worksheet(ws, data):
    """Write data to a worksheet with proper formatting."""
    # Determine the number of email columns needed
    max_emails = 1
    for row in data:
        emails = row.get("emails", [])
        max_emails = max(max_emails, len(emails))

    # Define the column order
    ordered_columns = ["name", "website"]
    for i in range(1, max_emails + 1):
        ordered_columns.append(f"email_{i}")
    ordered_columns.extend(["city", "state", "zip_code", "phone", "happy_hour"])

    # Write headers
    for col_num, column in enumerate(ordered_columns, 1):
        ws.cell(row=1, column=col_num, value=column)

    # Write data
    for row_num, row_data in enumerate(data, 2):
        # Prepare row data
        cleaned_data = {}

        # Process regular fields
        for column in ordered_columns:
            if column.startswith("email_"):
                continue  # Handle emails separately
            elif column == "happy_hour":
                cleaned_data[column] = "Yes"
            else:
                cleaned_data[column] = row_data.get(column, "")

        # Handle emails
        emails = row_data.get("emails", [])
        for i in range(min(max_emails, len(emails))):
            cleaned_data[f"email_{i+1}"] = emails[i]
        for i in range(len(emails), max_emails):
            cleaned_data[f"email_{i+1}"] = ""

        # Write to worksheet
        for col_num, column in enumerate(ordered_columns, 1):
            value = cleaned_data.get(column, "")
            try:
                ws.cell(row=row_num, column=col_num, value=value)
            except Exception as e:
                print(f"Error writing cell ({row_num},{col_num}): {str(e)}")
                ws.cell(row=row_num, column=col_num, value=str(value))

    # Auto-adjust column widths
    for col_num, column in enumerate(ordered_columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for cell in ws[column_letter]:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if max_length > 0 else 15
        ws.column_dimensions[column_letter].width = min(adjusted_width, 60)
